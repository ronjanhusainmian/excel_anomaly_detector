
from dataclasses import dataclass, field
from typing import List, Dict
import openpyxl
from openpyxl.utils import get_column_letter

from .signature import ParsedCell


@dataclass
class Block:
    sheet: str
    orientation: str 
    key: int 
    cells: List[ParsedCell] = field(default_factory=list)


def load_formula_cells(path: str) -> List[ParsedCell]:
    wb = openpyxl.load_workbook(path, data_only=False)
    cells: List[ParsedCell] = []

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    cells.append(
                        ParsedCell(
                            sheet=ws.title,
                            row=cell.row,
                            col=cell.column,
                            address=cell.coordinate,
                            formula=value,
                        )
                    )
    return cells


def _group_contiguous(indices: List[int]) -> List[List[int]]:
    if not indices:
        return []
    indices = sorted(indices)
    runs = [[indices[0]]]
    for n in indices[1:]:
        if n == runs[-1][-1] + 1:
            runs[-1].append(n)
        else:
            runs.append([n])
    return runs


def build_blocks(cells: List[ParsedCell], min_block_size: int = 3) -> List[Block]:
    blocks: List[Block] = []

    by_sheet_col: Dict[tuple, List[ParsedCell]] = {}
    by_sheet_row: Dict[tuple, List[ParsedCell]] = {}

    for c in cells:
        by_sheet_col.setdefault((c.sheet, c.col), []).append(c)
        by_sheet_row.setdefault((c.sheet, c.row), []).append(c)

    # Column-wise blocks (the more common "filled down a column" case)
    for (sheet, col), col_cells in by_sheet_col.items():
        rows = [c.row for c in col_cells]
        for run in _group_contiguous(rows):
            if len(run) >= min_block_size:
                run_set = set(run)
                block_cells = [c for c in col_cells if c.row in run_set]
                block_cells.sort(key=lambda c: c.row)
                blocks.append(Block(sheet=sheet, orientation="column", key=col, cells=block_cells))

    # Row-wise blocks (less common, but still possible)
    for (sheet, row), row_cells in by_sheet_row.items():
        cols = [c.col for c in row_cells]
        for run in _group_contiguous(cols):
            if len(run) >= min_block_size:
                run_set = set(run)
                block_cells = [c for c in row_cells if c.col in run_set]
                block_cells.sort(key=lambda c: c.col)
                blocks.append(Block(sheet=sheet, orientation="row", key=row, cells=block_cells))

    return blocks
